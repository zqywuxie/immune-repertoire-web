"""
CDR3 Export Service for the Immune Repertoire Analysis Web Application.
Provides functionality for exporting CDR3 shared lists and abundance matrices.

Export Formats:
1. CDR3_Shared_List.xlsx - Sample pair shared CDR3 sequences
2. {Chain}_Abundance_Union_Top100.xlsx - Top100 abundance matrix per chain
3. {Chain}_Abundance_Union_Full.xlsx - Full abundance matrix per chain
4. {Chain}_Top100_Analysis.xlsx - Top100 analysis with intersection matrix
"""
import io
import zipfile
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from collections import defaultdict

import pandas as pd
import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class CDR3ExportService:
    """
    Service for exporting CDR3 analysis results.
    Supports shared CDR3 lists and abundance matrices.
    """
    
    def __init__(self):
        """Initialize the CDR3 export service."""
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl is required for CDR3 export service")
    
    def export_shared_cdr3_pairs(
        self, 
        sample_data: Dict[str, pd.DataFrame], 
        top_n: int = 100
    ) -> pd.DataFrame:
        """
        Export shared CDR3 sequences between all sample pairs.
        
        Args:
            sample_data: Dictionary {sample_name: DataFrame with 'cdr3' and 'copy' columns}
            top_n: Number of top shared CDR3 to export per sample pair (0 = all)
        
        Returns:
            DataFrame containing shared CDR3 information for all sample pairs
        """
        sample_names = list(sample_data.keys())
        all_pairs_data = []
        
        for i, sample_a in enumerate(sample_names):
            for j, sample_b in enumerate(sample_names):
                if i >= j:  # Only process upper triangle (excluding diagonal)
                    continue
                
                df_a = sample_data[sample_a]
                df_b = sample_data[sample_b]
                
                # Get shared CDR3
                set_a = set(df_a['cdr3'].values)
                set_b = set(df_b['cdr3'].values)
                shared_cdr3 = set_a & set_b
                
                if not shared_cdr3:
                    continue
                
                # Build abundance dictionaries
                abundance_a = df_a.set_index('cdr3')['copy'].to_dict()
                abundance_b = df_b.set_index('cdr3')['copy'].to_dict()
                
                # Create records for each shared CDR3
                for cdr3 in shared_cdr3:
                    copy_a = abundance_a.get(cdr3, 0)
                    copy_b = abundance_b.get(cdr3, 0)
                    
                    all_pairs_data.append({
                        'Sample_A': sample_a,
                        'Sample_B': sample_b,
                        'CDR3': cdr3,
                        f'{sample_a}_Copy': copy_a,
                        f'{sample_b}_Copy': copy_b,
                        'Min_Copy': min(copy_a, copy_b),
                        'Max_Copy': max(copy_a, copy_b),
                        'Total_Copy': copy_a + copy_b
                    })
        
        if not all_pairs_data:
            return pd.DataFrame()
        
        # Create DataFrame and sort by total abundance
        result_df = pd.DataFrame(all_pairs_data)
        result_df = result_df.sort_values(['Sample_A', 'Sample_B', 'Total_Copy'], 
                                          ascending=[True, True, False])
        
        # If top_n specified, keep only top N per sample pair
        if top_n > 0:
            result_df = result_df.groupby(['Sample_A', 'Sample_B']).head(top_n)
        
        return result_df
    
    def create_shared_cdr3_excel(
        self, 
        shared_df: pd.DataFrame
    ) -> bytes:
        """
        Create Excel file with shared CDR3 list.
        Includes a summary sheet and individual sheets for each sample pair.
        
        Args:
            shared_df: DataFrame with shared CDR3 data
        
        Returns:
            Excel file as bytes
        """
        if shared_df.empty:
            raise ValueError("No shared CDR3 data to export")
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Create summary sheet
            summary_data = []
            for (sample_a, sample_b), group in shared_df.groupby(['Sample_A', 'Sample_B']):
                summary_data.append({
                    'Sample_A': sample_a,
                    'Sample_B': sample_b,
                    'Shared_CDR3_Count': len(group),
                    'Total_Abundance': group['Total_Copy'].sum(),
                    'Avg_Abundance': group['Total_Copy'].mean()
                })
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary sheet
            workbook = writer.book
            summary_sheet = workbook['Summary']
            self._format_excel_sheet(summary_sheet, is_header=True)
            
            # Create individual sheets for each sample pair
            for (sample_a, sample_b), group in shared_df.groupby(['Sample_A', 'Sample_B']):
                sheet_name = f"{sample_a}_vs_{sample_b}"[:31]  # Excel sheet name limit
                group_export = group[['CDR3', f'{sample_a}_Copy', f'{sample_b}_Copy', 
                                     'Min_Copy', 'Max_Copy', 'Total_Copy']]
                group_export.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Format sheet
                sheet = workbook[sheet_name]
                self._format_excel_sheet(sheet, is_header=True)
        
        output.seek(0)
        return output.read()
    
    def create_abundance_matrix_excel(
        self, 
        sample_data: Dict[str, pd.DataFrame],
        top_n: int = 0,
        chain_name: str = "All"
    ) -> bytes:
        """
        Create abundance matrix Excel file.
        Rows: CDR3 sequences, Columns: Sample names, Values: Copy numbers
        
        Args:
            sample_data: Dictionary of sample DataFrames
            top_n: Number of top CDR3 to include (0 = all)
            chain_name: Chain name for the file
        
        Returns:
            Excel file as bytes
        """
        # Collect all unique CDR3
        all_cdr3_abundance = defaultdict(lambda: defaultdict(int))
        
        for sample_name, df in sample_data.items():
            for _, row in df.iterrows():
                cdr3 = row['cdr3']
                copy = row['copy']
                all_cdr3_abundance[cdr3][sample_name] = copy
        
        # Calculate total abundance for each CDR3
        cdr3_totals = {cdr3: sum(samples.values()) 
                      for cdr3, samples in all_cdr3_abundance.items()}
        
        # Sort by total abundance
        sorted_cdr3 = sorted(cdr3_totals.keys(), key=lambda x: cdr3_totals[x], reverse=True)
        
        # Apply top_n filter if specified
        if top_n > 0:
            sorted_cdr3 = sorted_cdr3[:top_n]
        
        # Build matrix
        sample_names = list(sample_data.keys())
        matrix_data = []
        
        for cdr3 in sorted_cdr3:
            row = {'CDR3': cdr3}
            for sample_name in sample_names:
                row[sample_name] = all_cdr3_abundance[cdr3].get(sample_name, 0)
            row['Total'] = cdr3_totals[cdr3]
            matrix_data.append(row)
        
        matrix_df = pd.DataFrame(matrix_data)
        
        # Create Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            matrix_df.to_excel(writer, sheet_name=chain_name, index=False)
            
            # Format sheet
            workbook = writer.book
            sheet = workbook[chain_name]
            self._format_excel_sheet(sheet, is_header=True)
        
        output.seek(0)
        return output.read()
    
    def create_top100_analysis_excel(
        self,
        sample_data: Dict[str, pd.DataFrame],
        chain_name: str = "All"
    ) -> bytes:
        """
        Create Top100 analysis Excel file.
        First sheet: Intersection matrix of all samples' Top100
        Following sheets: Each sample's Top100 list
        
        Args:
            sample_data: Dictionary of sample DataFrames
            chain_name: Chain name for the file
        
        Returns:
            Excel file as bytes
        """
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Get Top100 for each sample
            sample_top100 = {}
            all_top100_cdr3 = set()
            
            for sample_name, df in sample_data.items():
                top100 = df.nlargest(100, 'copy')
                sample_top100[sample_name] = top100
                all_top100_cdr3.update(top100['cdr3'].values)
            
            # Create intersection matrix
            sample_names = list(sample_data.keys())
            matrix_data = []
            
            # Build abundance dict for all CDR3 in union of Top100
            for cdr3 in sorted(all_top100_cdr3):
                row = {'CDR3': cdr3}
                for sample_name in sample_names:
                    df = sample_data[sample_name]
                    cdr3_row = df[df['cdr3'] == cdr3]
                    row[sample_name] = cdr3_row['copy'].values[0] if len(cdr3_row) > 0 else 0
                matrix_data.append(row)
            
            # Sort by total abundance
            matrix_df = pd.DataFrame(matrix_data)
            matrix_df['Total'] = matrix_df[sample_names].sum(axis=1)
            matrix_df = matrix_df.sort_values('Total', ascending=False)
            
            # Write intersection matrix
            matrix_df.to_excel(writer, sheet_name='Intersection_Matrix', index=False)
            
            # Format intersection matrix sheet
            workbook = writer.book
            sheet = workbook['Intersection_Matrix']
            self._format_excel_sheet(sheet, is_header=True)
            
            # Write each sample's Top100
            for sample_name, top100_df in sample_top100.items():
                sheet_name = sample_name[:31]  # Excel sheet name limit
                export_df = top100_df[['cdr3', 'copy']].copy()
                export_df.columns = ['CDR3', 'Copy']
                export_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Format sheet
                sheet = workbook[sheet_name]
                self._format_excel_sheet(sheet, is_header=True)
        
        output.seek(0)
        return output.read()
    
    def _format_excel_sheet(self, sheet, is_header: bool = True):
        """
        Apply formatting to an Excel sheet.
        
        Args:
            sheet: openpyxl worksheet
            is_header: Whether to format the first row as header
        """
        # Format header row
        if is_header and sheet.max_row > 0:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, 
                                   min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Data rows
                    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def generate_complete_export_zip(
        self,
        sample_data: Dict[str, pd.DataFrame],
        include_summary: bool = True,
        top_n: int = 100
    ) -> bytes:
        """
        Generate complete export ZIP package containing:
        1. CDR3_Shared_List.xlsx - Sample pair shared CDR3 sequences
        2. Abundance_Union_Top100.xlsx - Top100 abundance matrix
        3. Abundance_Union_Full.xlsx - Full abundance matrix
        4. Top100_Analysis.xlsx - Top100 analysis with intersection matrix
        5. README.txt - Summary information
        
        Args:
            sample_data: Dictionary of sample DataFrames or chain-based nested dict
            include_summary: Whether to include README.txt
            top_n: Number of top CDR3 for shared list
        
        Returns:
            ZIP file as bytes
        """
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Check if data is chain-based (nested dict) or simple
            is_chain_based = any(isinstance(v, dict) for v in sample_data.values())
            
            if is_chain_based:
                # Chain-based export
                for chain, chain_samples in sample_data.items():
                    if not chain_samples:
                        continue
                    
                    # 1. Shared CDR3 list for this chain
                    shared_df = self.export_shared_cdr3_pairs(chain_samples, top_n=top_n)
                    if not shared_df.empty:
                        shared_excel = self.create_shared_cdr3_excel(shared_df)
                        zip_file.writestr(f'{chain}_CDR3_Shared_List.xlsx', shared_excel)
                    
                    # 2. Top100 abundance matrix
                    top100_matrix = self.create_abundance_matrix_excel(
                        chain_samples, top_n=100, chain_name=chain
                    )
                    zip_file.writestr(f'{chain}_Abundance_Union_Top100.xlsx', top100_matrix)
                    
                    # 3. Full abundance matrix
                    full_matrix = self.create_abundance_matrix_excel(
                        chain_samples, top_n=0, chain_name=chain
                    )
                    zip_file.writestr(f'{chain}_Abundance_Union_Full.xlsx', full_matrix)
                    
                    # 4. Top100 analysis
                    top100_analysis = self.create_top100_analysis_excel(
                        chain_samples, chain_name=chain
                    )
                    zip_file.writestr(f'{chain}_Top100_Analysis.xlsx', top100_analysis)
            
            else:
                # Simple export (all samples together)
                # 1. Shared CDR3 list
                shared_df = self.export_shared_cdr3_pairs(sample_data, top_n=top_n)
                if not shared_df.empty:
                    shared_excel = self.create_shared_cdr3_excel(shared_df)
                    zip_file.writestr('CDR3_Shared_List.xlsx', shared_excel)
                
                # 2. Top100 abundance matrix
                top100_matrix = self.create_abundance_matrix_excel(
                    sample_data, top_n=100, chain_name="All"
                )
                zip_file.writestr('Abundance_Union_Top100.xlsx', top100_matrix)
                
                # 3. Full abundance matrix
                full_matrix = self.create_abundance_matrix_excel(
                    sample_data, top_n=0, chain_name="All"
                )
                zip_file.writestr('Abundance_Union_Full.xlsx', full_matrix)
                
                # 4. Top100 analysis
                top100_analysis = self.create_top100_analysis_excel(
                    sample_data, chain_name="All"
                )
                zip_file.writestr('Top100_Analysis.xlsx', top100_analysis)
            
            # 5. Generate README
            if include_summary:
                readme_content = self._generate_readme(sample_data, is_chain_based)
                zip_file.writestr('README.txt', readme_content.encode('utf-8'))
        
        zip_buffer.seek(0)
        return zip_buffer.read()
    
    def _generate_readme(
        self, 
        sample_data: Dict, 
        is_chain_based: bool
    ) -> str:
        """Generate README content for the export package."""
        lines = [
            "=" * 60,
            "CDR3分析数据包说明文档",
            "=" * 60,
            f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "本压缩包包含以下CDR3分析结果文件：",
            "",
            "1. CDR3_Shared_List.xlsx",
            "   - 样本对共享CDR3列表",
            "   - 包含Summary汇总表和每个样本对的详细共享CDR3序列",
            "   - 用于查看两个样本之间的共享CDR3及其丰度",
            "",
            "2. {链类型}_Abundance_Union_Top100.xlsx",
            "   - 各链的丰度矩阵（Top100版本）",
            "   - 包含所有样本中出现过的CDR3序列",
            "   - 矩阵格式：行=CDR3序列，列=样本名，值=拷贝数",
            "   - 未出现的CDR3在样本中显示为0",
            "   - 按总丰度排序，只保留前100个",
            "",
            "3. {链类型}_Abundance_Union_Full.xlsx",
            "   - 各链的丰度矩阵（完整版本）",
            "   - 包含所有样本中出现过的所有CDR3序列（不限数量）",
            "   - 矩阵格式：行=CDR3序列，列=样本名，值=拷贝数",
            "   - 未出现的CDR3在样本中显示为0",
            "   - 按总丰度排序",
            "",
            "4. {链类型}_Top100_Analysis.xlsx",
            "   - 各链的Top100分析（合并文件）",
            "   - 第一个Sheet：交集矩阵",
            "     * 包含所有样本Top100的并集",
            "     * 矩阵格式：行=CDR3，列=样本，值=拷贝数",
            "   - 后续Sheet：每个样本的Top100",
            "     * 格式：左列=CDR3序列，右列=拷贝数",
            "     * 按拷贝数降序排列",
            "",
        ]
        
        # Add sample information
        if is_chain_based:
            lines.append("样本信息（按链分组）：")
            for chain, chain_samples in sample_data.items():
                if isinstance(chain_samples, dict):
                    lines.append(f"\n{chain}链:")
                    for sample_name, df in chain_samples.items():
                        unique_cdr3 = len(df)
                        total_reads = int(df['copy'].sum())
                        lines.append(f"  - {sample_name}: {unique_cdr3} unique CDR3, {total_reads} total reads")
        else:
            lines.append("样本信息：")
            for sample_name, df in sample_data.items():
                unique_cdr3 = len(df)
                total_reads = int(df['copy'].sum())
                lines.append(f"  - {sample_name}: {unique_cdr3} unique CDR3, {total_reads} total reads")
        
        lines.extend([
            "",
            "=" * 60
        ])
        
        return "\n".join(lines)


# Global service instance
_cdr3_export_service: Optional[CDR3ExportService] = None


def get_cdr3_export_service() -> CDR3ExportService:
    """Get the global CDR3 export service instance."""
    global _cdr3_export_service
    if _cdr3_export_service is None:
        _cdr3_export_service = CDR3ExportService()
    return _cdr3_export_service
