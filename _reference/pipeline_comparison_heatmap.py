#!/usr/bin/env python3
"""
Pipeline对比热图生成工具
用于对比三个不同pipeline (YXJ, DW, YPL) 的免疫组库分析结果

功能特点:
    - 同一样本的三个pipeline数据对比
    - 支持七条链 (IGH, IGK, IGL, TRA, TRB, TRD, TRG)
    - 自动处理不同pipeline的列名差异

列名映射:
    - YXJ: cdr3 + umi_counts
    - DW: CDR3(pep) + copy
    - YPL: CDR3(pep) + umi_number

使用方法:
    python pipeline_comparison_heatmap.py
"""

import os
import sys
import io
import zipfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import defaultdict
import json
from datetime import datetime

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
try:
    from matplotlib_venn import venn3
except ImportError:
    venn3 = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[!] openpyxl未安装,Excel导出功能不可用")

# 配置
BASE_DIR = Path("/workspace/data_shared/To_ZQY/260125")
OUTPUT_DIR = BASE_DIR / "pipeline_comparison_output"

# Pipeline配置
PIPELINE_CONFIG = {
    'YXJ': {
        'dir': BASE_DIR / 'YXJ',
        'cdr3_col': 'cdr3_aa',  # 使用氨基酸序列，与其他pipeline一致
        'copy_col': 'umi_counts',
        'file_pattern': '{sample}_{chain}.csv',  # DBY_IGH.csv
    },
    'DW': {
        'dir': BASE_DIR / 'DW',
        'cdr3_col': 'CDR3(pep)',
        'copy_col': 'copy',
        'file_pattern': 'SS03P_{sample}__{chain}.csv',  # SS03P_DBY__IGH.csv
    },
    'YPL': {
        'dir': BASE_DIR / 'YPL',
        'cdr3_col': 'CDR3(pep)',
        'copy_col': 'umi_number',
        'file_pattern': '{sample}-PER_{sample}_{chain}.csv',  # DBY-PER_DBY_IGH.csv
    }
}

# 样本列表
SAMPLES = ['DBY', 'GRD', 'SL', 'WXC', 'YCM']

# pipeline显示顺序
PIPELINE_ORDER = ['YXJ', 'DW', 'YPL']

# 七条链
CHAINS = ['IGH', 'IGK', 'IGL', 'TRA', 'TRB', 'TRD', 'TRG']

# 相似度指标
METRICS = ['expression_sharing', 'morisita_horn', 'cdr3_sharing', 'r2_inner', 'r2_outer', 'sorensen']

METRIC_NAMES = {
    'expression_sharing': 'Expression Sharing',
    'morisita_horn': 'Morisita-Horn Index',
    'cdr3_sharing': 'Unique CDR3 Sharing',
    'r2_inner': 'R² Inner',
    'r2_outer': 'R² Outer',
    'sorensen': 'Sorensen-Dice Index'
}

# 指标颜色方案
METRIC_COLOR_SCHEMES = {
    'r2_inner': 'Greens',
    'r2_outer': 'Purples',
    'cdr3_sharing': 'Reds',
    'expression_sharing': 'Blues',
    'morisita_horn': 'Oranges',
    'sorensen': 'YlGnBu'
}


def print_header(text: str):
    print("\n" + "=" * 60)
    print(f"  {text}")
    print("=" * 60)


def print_info(text: str):
    print(f"[INFO] {text}")


def print_success(text: str):
    print(f"[✓] {text}")


def print_warning(text: str):
    print(f"[!] {text}")


def print_error(text: str):
    print(f"[✗] {text}")


def format_similarity_value(value) -> str:
    """智能格式化相似度值"""
    import math
    
    if value is None:
        return "-"
    
    if isinstance(value, float):
        if math.isnan(value):
            return "-"
        if math.isinf(value):
            return "∞" if value > 0 else "-∞"
    
    try:
        val = float(value)
    except (ValueError, TypeError):
        return "-"
    
    if val == 0:
        return "0"
    
    if val == 1.0:
        return "1.000"
    
    abs_val = abs(val)
    
    if abs_val >= 0.01:
        return f"{val:.3f}"
    elif abs_val >= 0.001:
        return f"{val:.4f}"
    elif abs_val >= 0.0001:
        return f"{val:.5f}"
    else:
        return f"{val:.2e}"


def get_file_path(pipeline: str, sample: str, chain: str) -> Optional[Path]:
    """获取指定pipeline/样本/链的文件路径"""
    config = PIPELINE_CONFIG[pipeline]
    pattern = config['file_pattern']
    
    # 替换模式中的占位符
    filename = pattern.format(sample=sample, chain=chain)
    file_path = config['dir'] / filename
    
    if file_path.exists():
        return file_path
    
    return None


def normalize_cdr3(cdr3: str, pipeline: str) -> str:
    """
    标准化CDR3序列格式
    YXJ的cdr3_aa不含首尾保守氨基酸(如ASSFVGRPNEKLF)
    DW/YPL的CDR3(pep)含首尾(如CASSFVGRPNEKLFF)
    统一去除首尾保守氨基酸，只保留核心序列
    """
    if not isinstance(cdr3, str) or len(cdr3) < 3:
        return cdr3
    
    # DW和YPL的CDR3通常以C开头，以F/W结尾
    # 去除首尾保守氨基酸以统一格式
    if pipeline in ['DW', 'YPL']:
        # 去除开头的C（如果存在）
        if cdr3.startswith('C'):
            cdr3 = cdr3[1:]
        # 去除结尾的F或W（如果存在）
        if cdr3.endswith('F') or cdr3.endswith('W'):
            cdr3 = cdr3[:-1]
    
    return cdr3


def load_data(file_path: Path, cdr3_col: str, copy_col: str, pipeline: str = None) -> Optional[pd.DataFrame]:
    """加载并标准化数据"""
    try:
        # 检测分隔符
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            first_line = f.readline()
        sep = '\t' if '\t' in first_line else ','
        
        df = pd.read_csv(file_path, sep=sep, low_memory=False)
        
        if cdr3_col not in df.columns:
            print_warning(f"  列 '{cdr3_col}' 不存在于 {file_path.name}")
            return None
        if copy_col not in df.columns:
            print_warning(f"  列 '{copy_col}' 不存在于 {file_path.name}")
            return None
        
        # 标准化CDR3序列格式
        cdr3_series = df[cdr3_col].astype(str)
        if pipeline:
            cdr3_series = cdr3_series.apply(lambda x: normalize_cdr3(x, pipeline))
        
        # 标准化
        normalized_df = pd.DataFrame({
            'cdr3': cdr3_series,
            'copy': pd.to_numeric(df[copy_col], errors='coerce').fillna(0)
        })
        
        # 清理并聚合
        normalized_df = normalized_df.dropna(subset=['cdr3', 'copy'])
        normalized_df = normalized_df[normalized_df['cdr3'] != 'nan']
        normalized_df = normalized_df.groupby('cdr3', as_index=False)['copy'].sum()
        
        return normalized_df
        
    except Exception as e:
        print_error(f"  加载失败 {file_path.name}: {e}")
        return None


def calculate_pairwise_similarity(df1: pd.DataFrame, df2: pd.DataFrame, metric: str) -> float:
    """计算两个样本之间的相似度"""
    set1 = set(df1['cdr3'].values)
    set2 = set(df2['cdr3'].values)
    
    abundance1 = df1.set_index('cdr3')['copy'].to_dict()
    abundance2 = df2.set_index('cdr3')['copy'].to_dict()
    
    if metric == 'expression_sharing':
        shared_cdr3 = set1 & set2
        total_abundance_1 = sum(abundance1.values())
        
        if total_abundance_1 == 0:
            return 0.0
        
        shared_value = sum(min(abundance1.get(c, 0), abundance2.get(c, 0)) for c in shared_cdr3)
        return shared_value / total_abundance_1
    
    elif metric == 'morisita_horn':
        all_cdr3 = set1 | set2
        
        N_A = sum(abundance1.values())
        N_B = sum(abundance2.values())
        
        if N_A == 0 or N_B == 0:
            return 0.0
        
        all_list = sorted(all_cdr3)
        n_A = np.array([abundance1.get(c, 0) for c in all_list])
        n_B = np.array([abundance2.get(c, 0) for c in all_list])
        
        D_A = np.sum((n_A / N_A) ** 2)
        D_B = np.sum((n_B / N_B) ** 2)
        
        if D_A + D_B == 0:
            return 0.0
        
        numerator = 2 * np.sum(n_A * n_B)
        denominator = (D_A + D_B) * N_A * N_B
        
        if denominator > 0:
            return numerator / denominator
        else:
            return 0.0
    
    elif metric == 'cdr3_sharing':
        intersection = set1 & set2
        if len(set1) > 0:
            return len(intersection) / len(set1)
        else:
            return 0.0
    
    elif metric == 'r2_inner':
        shared_cdr3 = sorted(set1 & set2)
        
        if len(shared_cdr3) < 2:
            return 0.0
        
        x = np.array([abundance1[c] for c in shared_cdr3])
        y = np.array([abundance2[c] for c in shared_cdr3])
        
        std_x = np.std(x)
        std_y = np.std(y)
        
        if std_x == 0 and std_y == 0:
            return 1.0
        elif std_x == 0 or std_y == 0:
            return 0.0
        else:
            corr = np.corrcoef(x, y)[0, 1]
            return corr ** 2 if not np.isnan(corr) else 0.0
    
    elif metric == 'r2_outer':
        all_cdr3 = sorted(set1 | set2)
        
        if len(all_cdr3) < 2:
            return 0.0
        
        x = np.array([abundance1.get(c, 0) for c in all_cdr3])
        y = np.array([abundance2.get(c, 0) for c in all_cdr3])
        
        std_x = np.std(x)
        std_y = np.std(y)
        
        if std_x == 0 and std_y == 0:
            return 1.0
        elif std_x == 0 or std_y == 0:
            return 0.0
        else:
            corr = np.corrcoef(x, y)[0, 1]
            return corr ** 2 if not np.isnan(corr) else 0.0
    
    elif metric == 'sorensen':
        intersection = len(set1 & set2)
        size_sum = len(set1) + len(set2)
        return (2 * intersection) / size_sum if size_sum > 0 else 0.0
    
    return 0.0


def calculate_similarities(sample_data: Dict[str, pd.DataFrame]) -> Dict[str, np.ndarray]:
    """计算所有相似度指标"""
    sample_names = list(sample_data.keys())
    n = len(sample_names)
    
    results = {}
    symmetric_metrics = ['r2_inner', 'r2_outer', 'morisita_horn', 'sorensen']
    
    for metric in METRICS:
        matrix = np.zeros((n, n))
        
        for i, name_i in enumerate(sample_names):
            for j, name_j in enumerate(sample_names):
                if i == j:
                    matrix[i, j] = 1.0
                elif metric in symmetric_metrics and i > j:
                    matrix[i, j] = matrix[j, i]
                else:
                    similarity = calculate_pairwise_similarity(
                        sample_data[name_i],
                        sample_data[name_j],
                        metric
                    )
                    matrix[i, j] = similarity
                    if metric in symmetric_metrics:
                        matrix[j, i] = similarity
        
        results[metric] = matrix
    
    return results


def generate_heatmap(matrix: np.ndarray, sample_names: List[str], metric: str,
                     output_path: Path, title: str, show_values: bool = True,
                     axis_label: str = 'Pipeline'):
    """生成热图"""
    n_samples = len(sample_names)
    fig_size = max(8, min(22, n_samples * 0.95))
    fig, ax = plt.subplots(figsize=(fig_size, fig_size * 0.9))
    
    cmap = METRIC_COLOR_SCHEMES.get(metric, 'viridis')
    
    # 创建对角线遮罩
    mask = np.eye(len(matrix), dtype=bool)
    
    # 计算颜色范围
    matrix_copy = matrix.copy()
    np.fill_diagonal(matrix_copy, np.nan)
    vmin = np.nanmin(matrix_copy)
    vmax = np.nanmax(matrix_copy)
    
    if np.isnan(vmin) or np.isnan(vmax) or vmin == vmax:
        vmin = 0.0
        vmax = 1.0
    
    # 创建注释数组
    annot_data = None
    if show_values:
        annot_data = np.vectorize(format_similarity_value)(matrix)
    
    # 生成热图
    sns.heatmap(
        matrix,
        xticklabels=sample_names,
        yticklabels=sample_names,
        cmap=cmap,
        mask=mask,
        annot=annot_data if show_values else False,
        fmt="",
        square=True,
        vmin=vmin,
        vmax=vmax,
        linewidths=0.5,
        linecolor='#e0e0e0',
        cbar_kws={'label': 'Similarity', 'shrink': 0.8},
        ax=ax,
        annot_kws={'fontsize': 10}
    )
    
    ax.set_title(title, fontsize=14, fontweight='600', pad=15)
    ax.set_xlabel(axis_label, fontsize=12)
    ax.set_ylabel(axis_label, fontsize=12)
    
    tick_fontsize = 11 if n_samples <= 8 else 9
    ax.tick_params(axis='x', rotation=45, labelsize=tick_fontsize)
    ax.tick_params(axis='y', rotation=0, labelsize=tick_fontsize)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close(fig)


def save_matrix(matrix: np.ndarray, sample_names: List[str], output_path: Path):
    """保存相似度矩阵为CSV"""
    df = pd.DataFrame(matrix, index=sample_names, columns=sample_names)
    df.to_csv(output_path, encoding='utf-8-sig')


def export_shared_cdr3_pairs(sample_data: Dict[str, pd.DataFrame], top_n: int = 100) -> pd.DataFrame:
    """
    导出所有样本对之间的共享CDR3序列
    
    Args:
        sample_data: 样本数据字典 {sample_name: DataFrame with 'cdr3' and 'copy' columns}
        top_n: 每个样本对导出的top N共享CDR3数量
    
    Returns:
        包含所有样本对共享CDR3信息的DataFrame
    """
    sample_names = list(sample_data.keys())
    all_pairs_data = []
    
    for i, sample_a in enumerate(sample_names):
        for j, sample_b in enumerate(sample_names):
            if i >= j:  # 只处理上三角(不包括对角线)
                continue
            
            df_a = sample_data[sample_a]
            df_b = sample_data[sample_b]
            
            # 获取共享的CDR3
            set_a = set(df_a['cdr3'].values)
            set_b = set(df_b['cdr3'].values)
            shared_cdr3 = set_a & set_b
            
            if not shared_cdr3:
                continue
            
            # 构建丰度字典
            abundance_a = df_a.set_index('cdr3')['copy'].to_dict()
            abundance_b = df_b.set_index('cdr3')['copy'].to_dict()
            
            # 为每个共享CDR3创建记录
            for cdr3 in shared_cdr3:
                copy_a = abundance_a.get(cdr3, 0)
                copy_b = abundance_b.get(cdr3, 0)
                
                all_pairs_data.append({
                    'Sample_A': sample_a,
                    'Sample_B': sample_b,
                    'CDR3': cdr3,
                    f'{sample_a}_Copy': copy_a,
                    f'{sample_b}_Copy': copy_b,
                    'Min_Copy': min(copy_a, copy_b),
                    'Max_Copy': max(copy_a, copy_b),
                    'Total_Copy': copy_a + copy_b
                })
    
    if not all_pairs_data:
        return pd.DataFrame()
    
    # 创建DataFrame并按总丰度排序
    result_df = pd.DataFrame(all_pairs_data)
    result_df = result_df.sort_values(['Sample_A', 'Sample_B', 'Total_Copy'], 
                                      ascending=[True, True, False])
    
    # 如果指定了top_n,则每个样本对只保留top N
    if top_n > 0:
        result_df = result_df.groupby(['Sample_A', 'Sample_B']).head(top_n)
    
    return result_df


def create_shared_cdr3_excel(shared_df: pd.DataFrame, output_path: Path):
    """
    创建共享CDR3列表Excel文件
    包含Summary汇总表和每个样本对的详细Sheet
    
    Args:
        shared_df: 共享CDR3的DataFrame
        output_path: 输出文件路径
    """
    if not EXCEL_AVAILABLE:
        print_warning("openpyxl未安装,无法生成Excel文件")
        return
    
    if shared_df.empty:
        print_warning("没有共享CDR3数据可导出")
        return
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 创建Summary汇总表
        summary_data = []
        for (sample_a, sample_b), group in shared_df.groupby(['Sample_A', 'Sample_B']):
            summary_data.append({
                'Sample_A': sample_a,
                'Sample_B': sample_b,
                'Shared_CDR3_Count': len(group),
                'Total_Abundance': int(group['Total_Copy'].sum()),
                'Avg_Abundance': round(group['Total_Copy'].mean(), 2)
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # 格式化Summary表
        workbook = writer.book
        summary_sheet = workbook['Summary']
        format_excel_sheet(summary_sheet, is_header=True)
        
        # 为每个样本对创建单独的Sheet
        for (sample_a, sample_b), group in shared_df.groupby(['Sample_A', 'Sample_B']):
            sheet_name = f"{sample_a}_vs_{sample_b}"[:31]  # Excel sheet名称限制
            group_export = group[['CDR3', f'{sample_a}_Copy', f'{sample_b}_Copy', 
                                 'Min_Copy', 'Max_Copy', 'Total_Copy']].copy()
            group_export.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 格式化Sheet
            sheet = workbook[sheet_name]
            format_excel_sheet(sheet, is_header=True)


def create_abundance_matrix_excel(sample_data: Dict[str, pd.DataFrame], 
                                 output_path: Path,
                                 top_n: int = 0,
                                 chain_name: str = "All"):
    """
    创建丰度矩阵Excel文件
    行: CDR3序列, 列: 样本名称, 值: 拷贝数
    
    Args:
        sample_data: 样本数据字典
        output_path: 输出文件路径
        top_n: 保留top N个CDR3 (0=全部)
        chain_name: 链名称
    """
    if not EXCEL_AVAILABLE:
        print_warning("openpyxl未安装,无法生成Excel文件")
        return
    
    # 收集所有CDR3及其丰度
    all_cdr3_abundance = defaultdict(lambda: defaultdict(int))
    
    for sample_name, df in sample_data.items():
        for _, row in df.iterrows():
            cdr3 = row['cdr3']
            copy = row['copy']
            all_cdr3_abundance[cdr3][sample_name] = copy
    
    # 计算每个CDR3的总丰度
    cdr3_totals = {cdr3: sum(samples.values()) 
                  for cdr3, samples in all_cdr3_abundance.items()}
    
    # 按总丰度排序
    sorted_cdr3 = sorted(cdr3_totals.keys(), key=lambda x: cdr3_totals[x], reverse=True)
    
    # 应用top_n过滤
    if top_n > 0:
        sorted_cdr3 = sorted_cdr3[:top_n]
    
    # 构建矩阵
    sample_names = list(sample_data.keys())
    matrix_data = []
    
    for cdr3 in sorted_cdr3:
        row = {'CDR3': cdr3}
        for sample_name in sample_names:
            row[sample_name] = all_cdr3_abundance[cdr3].get(sample_name, 0)
        row['Total'] = cdr3_totals[cdr3]
        matrix_data.append(row)
    
    matrix_df = pd.DataFrame(matrix_data)
    
    # 创建Excel文件
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        matrix_df.to_excel(writer, sheet_name=chain_name, index=False)
        
        # 格式化Sheet
        workbook = writer.book
        sheet = workbook[chain_name]
        format_excel_sheet(sheet, is_header=True)


def create_top100_analysis_excel(sample_data: Dict[str, pd.DataFrame],
                                 output_path: Path,
                                 chain_name: str = "All"):
    """
    创建Top100分析Excel文件
    第一个Sheet: 所有样本Top100的交集矩阵
    后续Sheet: 每个样本的Top100列表
    
    Args:
        sample_data: 样本数据字典
        output_path: 输出文件路径
        chain_name: 链名称
    """
    if not EXCEL_AVAILABLE:
        print_warning("openpyxl未安装,无法生成Excel文件")
        return
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 获取每个样本的Top100
        sample_top100 = {}
        all_top100_cdr3 = set()
        
        for sample_name, df in sample_data.items():
            top100 = df.nlargest(100, 'copy')
            sample_top100[sample_name] = top100
            all_top100_cdr3.update(top100['cdr3'].values)
        
        # 创建交集矩阵
        sample_names = list(sample_data.keys())
        matrix_data = []
        
        # 为Top100并集中的所有CDR3构建丰度字典
        for cdr3 in sorted(all_top100_cdr3):
            row = {'CDR3': cdr3}
            for sample_name in sample_names:
                df = sample_data[sample_name]
                cdr3_row = df[df['cdr3'] == cdr3]
                row[sample_name] = int(cdr3_row['copy'].values[0]) if len(cdr3_row) > 0 else 0
            matrix_data.append(row)
        
        # 按总丰度排序
        matrix_df = pd.DataFrame(matrix_data)
        matrix_df['Total'] = matrix_df[sample_names].sum(axis=1)
        matrix_df = matrix_df.sort_values('Total', ascending=False)
        
        # 写入交集矩阵
        matrix_df.to_excel(writer, sheet_name='Intersection_Matrix', index=False)
        
        # 格式化交集矩阵Sheet
        workbook = writer.book
        sheet = workbook['Intersection_Matrix']
        format_excel_sheet(sheet, is_header=True)
        
        # 写入每个样本的Top100
        for sample_name, top100_df in sample_top100.items():
            sheet_name = sample_name[:31]  # Excel sheet名称限制
            export_df = top100_df[['cdr3', 'copy']].copy()
            export_df.columns = ['CDR3', 'Copy']
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 格式化Sheet
            sheet = workbook[sheet_name]
            format_excel_sheet(sheet, is_header=True)


def format_excel_sheet(sheet, is_header: bool = True):
    """
    格式化Excel表格
    
    Args:
        sheet: openpyxl worksheet对象
        is_header: 是否格式化第一行为表头
    """
    if not EXCEL_AVAILABLE:
        return
    
    # 格式化表头行
    if is_header and sheet.max_row > 0:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 自动调整列宽
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, 
                               min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # 数据行
                cell.alignment = Alignment(horizontal='left', vertical='center')


def export_cdr3_analysis_by_chain(chain_loaded_data: Dict[str, Dict[str, pd.DataFrame]],
                                  output_dir: Path,
                                  top_n: int = 100):
    """
    按链导出完整的CDR3分析数据
    为每条链生成4个文件:
    1. {Chain}_CDR3_Shared_List.xlsx - 样本对共享CDR3列表
    2. {Chain}_Abundance_Union_Top100.xlsx - Top100丰度矩阵
    3. {Chain}_Abundance_Union_Full.xlsx - 完整丰度矩阵
    4. {Chain}_Top100_Analysis.xlsx - Top100分析(交集矩阵+各样本Top100)
    
    Args:
        chain_loaded_data: 按链组织的样本数据 {chain: {sample_name: DataFrame}}
        output_dir: 输出目录
        top_n: 共享CDR3列表中每个样本对的top N
    """
    if not EXCEL_AVAILABLE:
        print_warning("openpyxl未安装,无法生成Excel文件")
        return
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    for chain, sample_data in chain_loaded_data.items():
        if not sample_data or len(sample_data) < 2:
            print_warning(f"链 {chain}: 样本数不足,跳过")
            continue
        
        print_info(f"\n处理链: {chain}")
        
        # 1. 共享CDR3列表
        shared_df = export_shared_cdr3_pairs(sample_data, top_n=top_n)
        if not shared_df.empty:
            shared_path = output_dir / f"{chain}_CDR3_Shared_List.xlsx"
            create_shared_cdr3_excel(shared_df, shared_path)
            print_success(f"  生成共享CDR3列表: {shared_path.name}")
        
        # 2. Top100丰度矩阵
        top100_path = output_dir / f"{chain}_Abundance_Union_Top100.xlsx"
        create_abundance_matrix_excel(sample_data, top100_path, top_n=100, chain_name=chain)
        print_success(f"  生成Top100丰度矩阵: {top100_path.name}")
        
        # 3. 完整丰度矩阵
        full_path = output_dir / f"{chain}_Abundance_Union_Full.xlsx"
        create_abundance_matrix_excel(sample_data, full_path, top_n=0, chain_name=chain)
        print_success(f"  生成完整丰度矩阵: {full_path.name}")
        
        # 4. Top100分析
        analysis_path = output_dir / f"{chain}_Top100_Analysis.xlsx"
        create_top100_analysis_excel(sample_data, analysis_path, chain_name=chain)
        print_success(f"  生成Top100分析: {analysis_path.name}")
    
    # 生成README文件
    readme_path = output_dir / "README.txt"
    generate_readme(readme_path, chain_loaded_data)
    print_success(f"\n生成说明文档: {readme_path.name}")


def generate_readme(output_path: Path, chain_loaded_data: Dict[str, Dict[str, pd.DataFrame]]):
    """
    生成README说明文档
    
    Args:
        output_path: 输出文件路径
        chain_loaded_data: 按链组织的样本数据
    """
    lines = [
        "CDR3分析数据包说明文档",
        "="*60,
        f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "本压缩包包含以下CDR3分析结果文件:",
        "",
        "1. {链类型}_CDR3_Shared_List.xlsx",
        "   - 样本对共享CDR3列表",
        "   - 包含Summary汇总表和每个样本对的详细共享CDR3序列",
        "   - 用于查看两个样本之间的共享CDR3及其丰度",
        "",
        "2. {链类型}_Abundance_Union_Top100.xlsx",
        "   - 各链的丰度矩阵(Top100版本)",
        "   - 包含所有样本中出现过的CDR3序列",
        "   - 矩阵格式: 行=CDR3序列, 列=样本名, 值=拷贝数",
        "   - 未出现的CDR3在样本中显示为0",
        "   - 按总丰度排序,只保留前100个",
        "",
        "3. {链类型}_Abundance_Union_Full.xlsx",
        "   - 各链的丰度矩阵(完整版本)",
        "   - 包含所有样本中出现过的所有CDR3序列(不限数量)",
        "   - 矩阵格式: 行=CDR3序列, 列=样本名, 值=拷贝数",
        "   - 未出现的CDR3在样本中显示为0",
        "   - 按总丰度排序",
        "",
        "4. {链类型}_Top100_Analysis.xlsx",
        "   - 各链的Top100分析(合并文件)",
        "   - 第一个Sheet: 交集矩阵",
        "     * 包含所有样本Top100的并集",
        "     * 矩阵格式: 行=CDR3, 列=样本, 值=拷贝数",
        "   - 后续Sheet: 每个样本的Top100",
        "     * 格式: 左列=CDR3序列, 右列=拷贝数",
        "     * 按拷贝数降序排列",
        "",
        "样本信息(按链分组):",
    ]
    
    for chain, sample_data in chain_loaded_data.items():
        if sample_data:
            lines.append(f"\n{chain}链:")
            for sample_name, df in sample_data.items():
                unique_cdr3 = len(df)
                total_reads = int(df['copy'].sum())
                lines.append(f"  - {sample_name}: {unique_cdr3} unique CDR3, {total_reads} total reads")
    
    lines.extend([
        "",
        "="*60
    ])
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))


def _venn3_subsets(set_a: set, set_b: set, set_c: set) -> Tuple[int, int, int, int, int, int, int]:
    """按matplotlib-venn约定计算三集合子集大小"""
    return (
        len(set_a - set_b - set_c),
        len(set_b - set_a - set_c),
        len((set_a & set_b) - set_c),
        len(set_c - set_a - set_b),
        len((set_a & set_c) - set_b),
        len((set_b & set_c) - set_a),
        len(set_a & set_b & set_c),
    )


def generate_venn_plots(pipeline_data: Dict[str, pd.DataFrame], sample: str, chain: str,
                        venn_output_path: Path, ratio_venn_output_path: Path):
    """生成3个pipeline的韦恩图和比例韦恩图"""
    if venn3 is None:
        print_warning("matplotlib-venn 未安装，跳过韦恩图绘制")
        return

    required_names = [f"{pipeline}_{sample}" for pipeline in PIPELINE_ORDER]
    if not all(name in pipeline_data for name in required_names):
        print_warning(f"  {sample}_{chain}: 缺少完整3个pipeline数据，跳过韦恩图")
        return

    set_yxj = set(pipeline_data[f"YXJ_{sample}"]['cdr3'].values)
    set_dw = set(pipeline_data[f"DW_{sample}"]['cdr3'].values)
    set_ypl = set(pipeline_data[f"YPL_{sample}"]['cdr3'].values)
    subsets = _venn3_subsets(set_yxj, set_dw, set_ypl)

    # 1) 计数韦恩图
    fig, ax = plt.subplots(figsize=(8.6, 8.0))
    venn3(
        subsets=subsets,
        set_labels=('YXJ', 'DW', 'YPL'),
        ax=ax,
    )
    ax.set_title(f"{sample}_{chain} - CDR3 Venn (Count)", fontsize=14, fontweight='600')
    plt.tight_layout()
    plt.savefig(venn_output_path, dpi=300, bbox_inches='tight')
    plt.close(fig)

    # 2) 比例韦恩图（标签显示占并集比例）
    fig, ax = plt.subplots(figsize=(8.6, 8.0))
    vobj = venn3(
        subsets=subsets,
        set_labels=('YXJ', 'DW', 'YPL'),
        ax=ax,
    )

    total_union = len(set_yxj | set_dw | set_ypl)
    if total_union > 0:
        subset_keys = ['100', '010', '110', '001', '101', '011', '111']
        for key, count in zip(subset_keys, subsets):
            label = vobj.get_label_by_id(key)
            if label is not None:
                ratio = count / total_union * 100
                label.set_text(f"{ratio:.1f}%\n({count})")

    ax.set_title(f"{sample}_{chain} - CDR3 Venn (Ratio)", fontsize=14, fontweight='600')
    plt.tight_layout()
    plt.savefig(ratio_venn_output_path, dpi=300, bbox_inches='tight')
    plt.close(fig)


def ordered_chain_labels(chain_loaded_data: Dict[str, pd.DataFrame], order_mode: str) -> List[str]:
    """生成单链多样本模式的显示顺序标签"""
    labels: List[str] = []

    if order_mode == 'by_sample':
        for sample in SAMPLES:
            for pipeline in PIPELINE_ORDER:
                label = f"{sample}_{pipeline}"
                if label in chain_loaded_data:
                    labels.append(label)
    else:
        for pipeline in PIPELINE_ORDER:
            for sample in SAMPLES:
                label = f"{sample}_{pipeline}"
                if label in chain_loaded_data:
                    labels.append(label)

    return labels


def run_pipeline_comparison():
    """运行pipeline对比分析"""
    print_header("Pipeline对比热图生成工具")
    print(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 创建输出目录结构
    output_base = OUTPUT_DIR / 'shared_analysis'
    heatmap_dir = output_base / 'heatmap'
    metric_dir = output_base / 'metric'
    
    heatmap_dir.mkdir(parents=True, exist_ok=True)
    metric_dir.mkdir(parents=True, exist_ok=True)

    single_heatmap_dir = heatmap_dir / 'single_sample'
    single_metric_dir = metric_dir / 'single_sample'
    multi_sample_heatmap_dir = heatmap_dir / 'multi_sample_by_sample'
    multi_sample_metric_dir = metric_dir / 'multi_sample_by_sample'
    multi_pipeline_heatmap_dir = heatmap_dir / 'multi_sample_by_pipeline'
    multi_pipeline_metric_dir = metric_dir / 'multi_sample_by_pipeline'
    venn_dir = output_base / 'venn' / 'single_sample'
    venn_ratio_dir = output_base / 'venn_ratio' / 'single_sample'

    for d in [
        single_heatmap_dir,
        single_metric_dir,
        multi_sample_heatmap_dir,
        multi_sample_metric_dir,
        multi_pipeline_heatmap_dir,
        multi_pipeline_metric_dir,
        venn_dir,
        venn_ratio_dir,
    ]:
        d.mkdir(parents=True, exist_ok=True)
    
    print_info(f"输出目录: {output_base}")

    # 缓存每条链下所有样本+pipeline数据，用于多样本合并热图
    chain_loaded_data: Dict[str, Dict[str, pd.DataFrame]] = {chain: {} for chain in CHAINS}
    
    # 遍历每个样本和每条链
    for sample in SAMPLES:
        print_header(f"处理样本: {sample}")
        
        for chain in CHAINS:
            print_info(f"\n处理链: {chain}")
            
            # 加载三个pipeline的数据
            pipeline_data = {}
            
            for pipeline in PIPELINE_ORDER:
                config = PIPELINE_CONFIG[pipeline]
                file_path = get_file_path(pipeline, sample, chain)
                
                if file_path is None:
                    print_warning(f"  {pipeline}: 文件不存在")
                    continue
                
                df = load_data(file_path, config['cdr3_col'], config['copy_col'], pipeline)
                
                if df is not None:
                    # 使用 pipeline_sample 作为名称
                    name = f"{pipeline}_{sample}"
                    pipeline_data[name] = df
                    chain_loaded_data[chain][f"{sample}_{pipeline}"] = df
                    print_success(f"  {pipeline}: 加载 {len(df)} 条CDR3")
            
            if len(pipeline_data) < 2:
                print_warning(f"  样本数不足，跳过 {sample}_{chain}")
                continue
            
            # 创建样本+链的子目录
            sample_chain = f"{sample}_{chain}"
            chain_heatmap_dir = single_heatmap_dir / sample_chain
            chain_metric_dir = single_metric_dir / sample_chain
            chain_venn_dir = venn_dir / sample_chain
            chain_venn_ratio_dir = venn_ratio_dir / sample_chain
            chain_heatmap_dir.mkdir(parents=True, exist_ok=True)
            chain_metric_dir.mkdir(parents=True, exist_ok=True)
            chain_venn_dir.mkdir(parents=True, exist_ok=True)
            chain_venn_ratio_dir.mkdir(parents=True, exist_ok=True)
            
            # 计算相似度
            similarities = calculate_similarities(pipeline_data)
            sample_names = list(pipeline_data.keys())
            
            # 生成热图和保存矩阵
            for metric in METRICS:
                title = f"{sample}_{chain} - {METRIC_NAMES[metric]}"
                
                # 热图
                heatmap_path = chain_heatmap_dir / f"{metric}_pipeline_comparison.png"
                generate_heatmap(
                    similarities[metric],
                    sample_names,
                    metric,
                    heatmap_path,
                    title,
                    axis_label='Pipeline',
                )
                
                # 矩阵CSV
                matrix_path = chain_metric_dir / f"{metric}_pipeline_comparison.csv"
                save_matrix(similarities[metric], sample_names, matrix_path)

            # 生成3个pipeline韦恩图和比例韦恩图
            generate_venn_plots(
                pipeline_data=pipeline_data,
                sample=sample,
                chain=chain,
                venn_output_path=chain_venn_dir / 'cdr3_venn.png',
                ratio_venn_output_path=chain_venn_ratio_dir / 'cdr3_venn_ratio.png',
            )
            
            print_success(f"  已生成 {sample_chain} 的对比热图")

    # 2) 单链多样本合并：按样本排序（sample1_YXJ, sample1_DW, sample1_YPL, ...）
    print_header('生成单链多样本热图（按样本排序）')
    for chain in CHAINS:
        chain_data = chain_loaded_data[chain]
        ordered_labels = ordered_chain_labels(chain_data, 'by_sample')
        if len(ordered_labels) < 2:
            print_warning(f"链 {chain}: 数据不足，跳过按样本排序热图")
            continue

        ordered_data = {label: chain_data[label] for label in ordered_labels}
        similarities = calculate_similarities(ordered_data)

        chain_heatmap_dir = multi_sample_heatmap_dir / chain
        chain_metric_dir = multi_sample_metric_dir / chain
        chain_heatmap_dir.mkdir(parents=True, exist_ok=True)
        chain_metric_dir.mkdir(parents=True, exist_ok=True)

        for metric in METRICS:
            title = f"{chain} - {METRIC_NAMES[metric]} (Order: by sample)"
            heatmap_path = chain_heatmap_dir / f"{metric}_by_sample.png"
            matrix_path = chain_metric_dir / f"{metric}_by_sample.csv"

            generate_heatmap(
                similarities[metric],
                ordered_labels,
                metric,
                heatmap_path,
                title,
                axis_label='Sample-Pipeline',
            )
            save_matrix(similarities[metric], ordered_labels, matrix_path)

        print_success(f"  已生成 {chain} 按样本排序热图")

    # 3) 单链多样本合并：按pipeline排序（YXJ全样本, DW全样本, YPL全样本）
    print_header('生成单链多样本热图（按pipeline排序）')
    for chain in CHAINS:
        chain_data = chain_loaded_data[chain]
        ordered_labels = ordered_chain_labels(chain_data, 'by_pipeline')
        if len(ordered_labels) < 2:
            print_warning(f"链 {chain}: 数据不足，跳过按pipeline排序热图")
            continue

        ordered_data = {label: chain_data[label] for label in ordered_labels}
        similarities = calculate_similarities(ordered_data)

        chain_heatmap_dir = multi_pipeline_heatmap_dir / chain
        chain_metric_dir = multi_pipeline_metric_dir / chain
        chain_heatmap_dir.mkdir(parents=True, exist_ok=True)
        chain_metric_dir.mkdir(parents=True, exist_ok=True)

        for metric in METRICS:
            title = f"{chain} - {METRIC_NAMES[metric]} (Order: by pipeline)"
            heatmap_path = chain_heatmap_dir / f"{metric}_by_pipeline.png"
            matrix_path = chain_metric_dir / f"{metric}_by_pipeline.csv"

            generate_heatmap(
                similarities[metric],
                ordered_labels,
                metric,
                heatmap_path,
                title,
                axis_label='Sample-Pipeline',
            )
            save_matrix(similarities[metric], ordered_labels, matrix_path)

        print_success(f"  已生成 {chain} 按pipeline排序热图")
    
    # 保存元数据
    metadata = {
        'timestamp': datetime.now().isoformat(),
        'samples': SAMPLES,
        'chains': CHAINS,
        'pipelines': PIPELINE_ORDER,
        'metrics': METRICS,
        'output_modes': {
            'single_sample': '单样本下三个pipeline比较',
            'multi_sample_by_sample': '单链多样本按样本排序',
            'multi_sample_by_pipeline': '单链多样本按pipeline排序',
            'venn_single_sample': '单样本-单链 3个pipeline韦恩图（计数）',
            'venn_ratio_single_sample': '单样本-单链 3个pipeline比例韦恩图',
        },
        'pipeline_config': {
            k: {'cdr3_col': v['cdr3_col'], 'copy_col': v['copy_col']} 
            for k, v in PIPELINE_CONFIG.items()
        }
    }
    
    metadata_path = output_base / 'metadata.json'
    with open(metadata_path, 'w', encoding='utf-8') as f:
        json.dump(metadata, f, indent=2, ensure_ascii=False)
    
    # 4) 导出CDR3分析数据（按链组织）
    print_header('导出CDR3分析数据')
    cdr3_analysis_dir = output_base / 'cdr3_analysis'
    
    # 重新组织数据：从 {chain: {sample_pipeline: df}} 转换为 {chain: {sample: df}}
    # 这里我们需要为每条链合并同一样本的不同pipeline数据
    chain_sample_data = {}
    for chain in CHAINS:
        chain_sample_data[chain] = {}
        for sample in SAMPLES:
            # 收集该样本在该链下所有pipeline的数据
            sample_dfs = []
            for pipeline in PIPELINE_ORDER:
                key = f"{sample}_{pipeline}"
                if key in chain_loaded_data[chain]:
                    sample_dfs.append(chain_loaded_data[chain][key])
            
            # 如果有数据，合并（去重并求和）
            if sample_dfs:
                combined_df = pd.concat(sample_dfs, ignore_index=True)
                combined_df = combined_df.groupby('cdr3', as_index=False)['copy'].sum()
                chain_sample_data[chain][sample] = combined_df
    
    # 导出CDR3分析
    export_cdr3_analysis_by_chain(chain_sample_data, cdr3_analysis_dir, top_n=100)
    
    print_header("完成！")
    print_success(f"输出目录: {output_base}")
    print("\n生成的文件结构:")
    print(f"  {output_base.name}/")
    print(f"    ├── heatmap/")
    print(f"    │   ├── single_sample/  (单样本三pipeline热图)")
    print(f"    │   │   └── DBY_IGH/ ...")
    print(f"    │   ├── multi_sample_by_sample/  (单链多样本按样本排序)")
    print(f"    │   │   └── IGH/ ...")
    print(f"    │   └── multi_sample_by_pipeline/  (单链多样本按pipeline排序)")
    print(f"    ├── metric/")
    print(f"    │   ├── single_sample/")
    print(f"    │   ├── multi_sample_by_sample/")
    print(f"    │   └── multi_sample_by_pipeline/")
    print(f"    ├── venn/")
    print(f"    │   └── single_sample/  (3个pipeline计数韦恩图)")
    print(f"    ├── venn_ratio/")
    print(f"    │   └── single_sample/  (3个pipeline比例韦恩图)")
    print(f"    ├── cdr3_analysis/  (CDR3分析数据)")
    print(f"    │   ├── {CHAINS[0]}_CDR3_Shared_List.xlsx")
    print(f"    │   ├── {CHAINS[0]}_Abundance_Union_Top100.xlsx")
    print(f"    │   ├── {CHAINS[0]}_Abundance_Union_Full.xlsx")
    print(f"    │   ├── {CHAINS[0]}_Top100_Analysis.xlsx")
    print(f"    │   ├── ... (其他链的文件)")
    print(f"    │   └── README.txt")
    print(f"    └── metadata.json")


if __name__ == '__main__':
    run_pipeline_comparison()
